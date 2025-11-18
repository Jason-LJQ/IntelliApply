import re
import functools
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import os
import shutil
import tempfile
from datetime import datetime

from intelliapply.utils.print_utils import print_, print_results
from intelliapply.utils.string_utils import normalize_company_name, get_abbreviation_lower, cleaned_string
from intelliapply.config.credential import EXCEL_FILE_PATH
from intelliapply.config.prompt import ALL_FIELDS


def sync(func):
    """Decorator: auto call _sync_data() before function execution."""

    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        if hasattr(self, "_sync_data"):
            self._sync_data()
        return func(self, *args, **kwargs)

    return wrapper


def save(func):
    """Decorator: auto call _save_data() asynchronously after function execution."""

    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        result = func(self, *args, **kwargs)
        if hasattr(self, "_save_data"):
            # Start async save in background thread
            save_thread = threading.Thread(target=self._save_data, daemon=True)
            save_thread.start()
        return result

    return wrapper


class ExcelManager:
    """
    Object-oriented Excel manager with intelligent caching and conflict detection.
    Loads all data including cell colors during DataFrame read, avoiding repeated disk I/O.
    """

    # Status color definitions
    STATUS_COLORS = {
        'REJECTED': 'FFFF0000',  # Red
        'PROCESSING': 'FFFFFF00',  # Yellow
        'OFFER': 'FF00FF00'  # Green
    }

    STATUS_SYMBOLS = {
        'REJECTED': '  ⨉  ',
        'PROCESSING': '  →  ',
        'OFFER': '  ✔  ',
        'NONE': ''
    }

    def __init__(self, file_path=None):
        """
        Initialize ExcelManager with file path.

        Args:
            file_path: Path to the Excel file (defaults to EXCEL_FILE_PATH from config)
        """
        self.file_path = file_path or EXCEL_FILE_PATH
        if not self.validate_excel_file():
            print_("Excel file is invalid. Please check the file and try again.", "RED")
            exit(1)
        self._cached_df = None
        self._cached_workbook = None
        self._last_mtime = 0.0

    def _color_to_status(self, color_rgb):
        """
        Convert RGB color to internal status identifier.

        Args:
            color_rgb: RGB color code from openpyxl cell

        Returns:
            Status string ('REJECTED', 'PROCESSING', 'OFFER', or 'NONE')
        """
        if color_rgb == self.STATUS_COLORS['REJECTED']:
            return 'REJECTED'
        elif color_rgb == self.STATUS_COLORS['PROCESSING']:
            return 'PROCESSING'
        elif color_rgb == self.STATUS_COLORS['OFFER']:
            return 'OFFER'
        else:
            return 'NONE'

    def _status_to_symbol(self, status):
        """
        Convert internal status to display symbol.

        Args:
            status: Internal status string

        Returns:
            Display symbol string
        """
        return self.STATUS_SYMBOLS.get(status, '')

    def _get_current_mtime(self):
        """Get current modification time of Excel file."""
        try:
            return os.path.getmtime(self.file_path)
        except Exception:
            return 0.0

    def _sync_data(self):
        """
        Synchronize cache with Excel file if needed.
        Automatically decides whether to invalidate cache based on file mtime.
        No return value - updates internal cache state only.
        """
        current_mtime = self._get_current_mtime()

        # Check if cache is valid (within 2 seconds tolerance)
        if self._cached_df is not None and abs(current_mtime - self._last_mtime) < 2:
            return

        # Cache invalid, reload from disk
        self.invalidate_cache()
        print_(f"Reloading Excel data from file", "YELLOW")
        try:
            # Load data with pandas
            df = pd.read_excel(self.file_path)

            # Load workbook with openpyxl to read cell colors
            wb = load_workbook(filename=self.file_path)
            ws = wb.active

            # Get Status column index
            headers = {cell.value: cell.column for cell in ws[1]}

            if 'Status' not in headers:
                # If no Status column, add empty _internal_status column
                raise ValueError("'Status' column not found in the Excel file.")
            else:
                status_col_idx = headers['Status']

            # Read status colors for all data rows
            status_list = []
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                cell = ws.cell(row=row_idx, column=status_col_idx)
                color_rgb = cell.fill.start_color.rgb if cell.fill else None
                status = self._color_to_status(color_rgb)
                status_list.append(status)

            # Add internal status column to DataFrame
            df['_internal_status'] = status_list

            # Update cache (keep workbook open for future writes)
            self._cached_df = df
            self._cached_workbook = wb
            self._last_mtime = current_mtime

        except Exception as e:
            print_(f"Error loading Excel data: {str(e)}", "RED")
            self.invalidate_cache()

    def _save_data(self):
        """
        Save cached workbook to disk and update mtime.
        Reports errors if save fails.
        """
        if self._cached_workbook is not None:
            try:
                self._cached_workbook.save(self.file_path)
                self._last_mtime = self._get_current_mtime()
            except Exception as e:
                print_(f"Error saving workbook: {str(e)}", "RED")
                raise

    def invalidate_cache(self):
        """Public method to force clear Excel data cache."""
        if self._cached_df is not None:
            self._cached_df = None
            self._last_mtime = 0.0

        # Close and clear cached workbook
        if self._cached_workbook is not None:
            self._cached_workbook.close()
            self._cached_workbook = None

            print_("Excel data will be reloaded from disk next time.", "YELLOW")

    def _check_for_write_conflict(self):
        """
        Check if file has been modified externally since last read.
        Warns user and requests confirmation if conflict detected.

        Returns:
            True if safe to write (no conflict or user confirmed), False otherwise
        """
        if self._last_mtime == 0.0:
            # No previous read, safe to write
            return True

        current_mtime = self._get_current_mtime()

        # Check if file modified externally (2 second tolerance)
        if abs(current_mtime - self._last_mtime) > 2:
            print_("WARNING: Excel file has been modified externally since last read!", "YELLOW")
            print_(f"Last read: {datetime.fromtimestamp(self._last_mtime).strftime('%Y-%m-%d %H:%M:%S')}", "YELLOW")
            print_(f"Current: {datetime.fromtimestamp(current_mtime).strftime('%Y-%m-%d %H:%M:%S')}", "YELLOW")

            confirm = input(
                print_("Continue with write operation? This may overwrite external changes. (y/yes to confirm): ",
                       color="BLUE", return_text=True)).lower()

            if confirm not in ['y', 'yes']:
                print_("Write operation cancelled to prevent data loss.", "RED")
                return False

        return True

    def validate_excel_file(self):
        """
        Validates the Excel file:
        1. Check if file exists, if not, ask user to create
        2. Check if all required columns exist, if not, backup the file and create a new one

        Returns:
            bool: True if validation passes or issues are fixed, False if validation fails
        """

        def create_new_excel():
            wb = Workbook()
            ws = wb.active
            for col, field in enumerate(['Status'] + ALL_FIELDS, 1):
                ws.cell(row=1, column=col, value=field)
            wb.save(self.file_path)
            print_(f"Created new Excel file at {self.file_path}", "GREEN")

        try:
            # Check if file exists
            if not os.path.exists(self.file_path):
                dir_path = os.path.dirname(self.file_path)
                if not os.path.exists(dir_path):
                    os.makedirs(dir_path)

                confirm = input(
                    print_(f"Excel file not found at {self.file_path}. Create it? (y/Y to confirm): ", color="BLUE",
                           return_text=True)).lower()
                if confirm == 'y':
                    # Create new Excel file with required columns
                    create_new_excel()
                    return True
                else:
                    print_("Excel file creation cancelled.", "RED")
                    return False

            # Check if all required columns exist
            wb = load_workbook(filename=self.file_path)
            ws = wb.active
            existing_headers = [cell.value for cell in ws[1]]
            missing_headers = [header for header in ['Status'] + ALL_FIELDS if header not in existing_headers]
            wb.close()

            if missing_headers:
                # Confirm to backup and create new file
                confirm = input(
                    print_(f"Missing columns: {missing_headers}. Backup and create new file? (y/Y to confirm): ",
                           color="BLUE", return_text=True)).lower()
                if confirm != 'y':
                    print_("Validation failed.", "RED")
                    return False

                # Backup existing file using shutil
                backup_path = f"{os.path.splitext(self.file_path)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_bak.xlsx"
                shutil.move(self.file_path, backup_path)
                print_(f"Backed up existing file to {backup_path}", "YELLOW")

                # Create new file with correct headers
                create_new_excel()
                return True

            return True

        except Exception as e:
            print_(f"Error validating Excel file: {str(e)}", "RED")
            return False

    @sync
    def search_applications(self, search_term="", index=-1):
        """
        Search for both company and job title matches using cached DataFrame.

        Args:
            search_term: The search term to match against company and job title
            index: If >= 0, directly return the record at this index (1-indexed, including header row)

        Returns:
            List of matching records
        """
        try:
            df = self._cached_df

            if df.empty:
                return []

            def applied_date(row):
                raw = row.get('Applied Date', '')
                return '' if pd.isna(raw) else raw

            # If index is provided, directly return that record
            if index >= 2:  # Index 1 is header
                if index - 2 < len(df):  # Convert to 0-indexed for DataFrame
                    row = df.iloc[index - 2]

                    result = [{
                        'Company': row['Company'],
                        'Location': row['Location'],
                        'Job Title': row['Job Title'],
                        'Applied Date': applied_date(row),
                        'Processed Date': row.get('Processed Date', ''),
                        'Result Date': row.get('Result Date', ''),
                        'status': self._status_to_symbol(row.get('_internal_status', 'NONE')),
                        'row_index': index,
                    }]
                    return result
                else:
                    print_(f"Index {index} is out of range.", "RED")
                    return []

            # 1. Prepare search term variations
            search_term_clean_lower = cleaned_string(search_term).lower()
            norm_keyword = normalize_company_name(search_term)
            abbr_keyword = get_abbreviation_lower(norm_keyword)

            # Ensure target columns are string type
            df['Company'] = df['Company'].astype(str).fillna('')
            df['Job Title'] = df['Job Title'].astype(str).fillna('')

            # 2. Create helper columns using vectorized `apply`
            df['norm_company'] = df['Company'].apply(normalize_company_name)
            df['abbr_company'] = df['Company'].apply(get_abbreviation_lower)
            df['clean_job_title'] = df['Job Title'].apply(cleaned_string).str.lower()

            # Create a regex pattern for word-level matching
            job_title_pattern = r'\b' + re.escape(search_term_clean_lower) + r'\b'

            # 3. Build boolean masks that perfectly replicate the original function's logic

            # Mask 1: Direct, Prefix, and Job Title matches
            m_base = (
                    (df['norm_company'] == norm_keyword) |
                    (df['norm_company'].str.startswith(norm_keyword, na=False)) |
                    (df['clean_job_title'].str.contains(job_title_pattern, na=False, regex=True))
            )

            # Mask 2: Handles "om" matching "Old Mission"
            m_abbr_target = (df['abbr_company'] == norm_keyword)

            # Mask 3: Handles "GSK" matching "GlaxoSmithKline"
            m_abbr_keyword_vs_abbr_target = pd.Series([False] * len(df), index=df.index)
            if len(abbr_keyword) > 1:
                m_abbr_keyword_vs_abbr_target = (df['abbr_company'] == abbr_keyword)

            # 4. Combine all masks using logical OR
            final_mask = m_base | m_abbr_target | m_abbr_keyword_vs_abbr_target

            # 5. Filter the DataFrame and format results
            matched_df = df[final_mask]

            if matched_df.empty:
                return []

            matches = []
            for idx, row in matched_df.iterrows():
                excel_row_index = idx + 2  # Convert 0-based DataFrame index to 1-based Excel row index
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row.to_dict()),
                    'Processed Date': row.get('Processed Date', ''),
                    'Result Date': row.get('Result Date', ''),
                    'status': self._status_to_symbol(row.get('_internal_status', 'NONE')),
                    'row_index': excel_row_index,
                })

            return matches

        except Exception as e:
            print_(f"Error during search: {str(e)}", "RED")
            return []

    @sync
    def check_duplicate_entry(self, new_data=None):
        """
        Check if the exact same job entry already exists in the Excel file.
        Returns matching row if duplicate found, None otherwise.
        """
        if new_data is None:
            return None
        try:
            # Create a deep copy to avoid polluting cache
            df = self._cached_df.copy()

            if df.empty:
                return None

            # Convert all columns to string for comparison
            for col in df.columns:
                if col != '_internal_status':  # Skip internal column
                    df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

            # Check each row for exact match
            for _, row in df.iterrows():
                all_fields_match = True
                for field in ['Company', 'Job Title']:
                    if field in row and field in new_data:
                        if str(row[field]).strip() != str(new_data[field]).strip():
                            all_fields_match = False
                            break
                if all_fields_match:
                    return row
            return None

        except Exception as e:
            print_(f"Error checking for duplicates: {str(e)}", "RED")
            return None

    @sync
    def summary(self):
        """
        Print a summary of job applications including:
        - Total number of applications
        - Number of rejections, processing, and offers
        - Rejection, processing, and offer rates
        """
        try:
            df = self._cached_df

            if df.empty:
                print_("No applications to summarize.", "RED")
                return

            total_applications = len(df)

            # Count status from _internal_status column
            rejections = len(df[df['_internal_status'] == 'REJECTED'])
            processing = len(df[df['_internal_status'] == 'PROCESSING'])
            offers = len(df[df['_internal_status'] == 'OFFER'])

            # Calculate percentages
            rejection_rate = (rejections / total_applications * 100) if total_applications > 0 else 0
            processing_rate = (processing / total_applications * 100) if total_applications > 0 else 0
            offer_rate = (offers / processing * 100) if processing > 0 else 0

            # Print summary with color formatting
            print_(f"\nApplication Summary:")
            print(f"Total Applications: {total_applications}")
            print(f"Rejected: {rejections}")
            print(f"Processing: {processing}")
            print(f"Offers: {offers}")
            print(f"Rejection Rate: {rejection_rate:.1f}%")
            print(f"Processing Rate: {processing_rate:.1f}%")
            print(f"Offer Rate: {offer_rate:.1f}% (offers/processing)")

        except Exception as e:
            print_(f"Error generating summary: {str(e)}", "RED")

    @sync
    def show_last_row(self, delete=False):
        """
        Show last row's data. If delete is True, delete the last row after user confirmation.
        """
        try:
            df = self._cached_df

            if df.empty:
                print_("No data to show.", "RED")
                return

            # Get last row data
            results = self.search_applications(index=len(df) + 1)  # +1 for header row
            print_results(results)

            if delete:
                # Ask for user confirmation
                confirm = input(print_("Delete this row? (y/Y to confirm, any other key to cancel): ", color="BLUE",
                                       return_text=True)).lower()
                if confirm == 'y':
                    self._delete_last_row()
                else:
                    print_(f"Deletion cancelled.", "RED")

        except Exception as e:
            print_(f"Error processing last row: \n{str(e)}", "RED")

    @sync
    @save
    def _delete_last_row(self):
        """
        Internal method to delete the last row from workbook.
        Synchronously updates cached DataFrame.
        """
        try:
            # Check for write conflict before deleting
            if not self._check_for_write_conflict():
                return

            sheet = self._cached_workbook.active
            last_row = sheet.max_row
            sheet.delete_rows(last_row)

            # Update cached DataFrame: drop last row
            if self._cached_df is not None and len(self._cached_df) > 0:
                self._cached_df.drop(self._cached_df.index[-1], inplace=True)
                self._cached_df.reset_index(drop=True, inplace=True)  # Reset index to avoid gaps

            print_(f"Last row deleted successfully.", "GREEN")

        except Exception as e:
            print_(f"Error deleting last row: {str(e)}", "RED")
            raise

    def open_excel_file(self):
        """
        Opens the Excel file using the default application based on the operating system.
        Returns True if successful, False otherwise.
        """
        try:
            import subprocess
            import platform
            system = platform.system()
            if system == 'Darwin':  # macOS
                subprocess.run(['open', self.file_path])
            elif system == 'Windows':
                subprocess.run(['start', '', self.file_path], shell=True)
            else:  # Linux and other OS
                subprocess.run(['xdg-open', self.file_path])
            print_(f"Excel file opened successfully.", "GREEN")
        except Exception as e:
            print_(f"Error opening Excel file: {str(e)}", "RED")
            return False
        return True

    @sync
    @save
    def append_data_to_excel(self, data=None):
        """
        Appends a list of dictionaries to the end of the Excel file.
        Uses cached workbook if available.
        Synchronously updates cached DataFrame.
        """
        if data is None:
            return

        try:
            # Check for write conflict before appending
            if not self._check_for_write_conflict():
                return

            # Use cached workbook or load new one
            if self._cached_workbook is None:
                workbook = load_workbook(filename=self.file_path)
                self._cached_workbook = workbook
            else:
                workbook = self._cached_workbook

            sheet = workbook.active

            # Find the last row with data (or the header row if the sheet is empty)
            last_row = sheet.max_row if sheet.max_row > 1 else 1

            # Get the headers and their column indices
            headers = {cell.value: cell.column for cell in sheet[1]}

            # List of all possible columns
            all_columns = ALL_FIELDS

            # Add any missing columns
            next_column = len(headers) + 1
            for column in all_columns:
                if column not in headers:
                    headers[column] = next_column
                    sheet.cell(row=1, column=next_column, value=column)
                    next_column += 1

            # Append the new data directly to the last row
            for row_data in data:
                last_row += 1  # Move to the next row for each new record

                # Write data to the corresponding columns
                for column in all_columns:
                    if column != 'Status':  # Skip Status column as it's handled separately
                        sheet.cell(row=last_row, column=headers[column],
                                   value=row_data.get(column, ''))

            new_df = pd.DataFrame(data)
            new_df['_internal_status'] = 'NONE'  # Initialize status for new rows
            self._cached_df = pd.concat([self._cached_df, new_df], ignore_index=True)

        except Exception as e:
            print_(f"Error appending data to Excel: {str(e)}")

    @sync
    @save
    def _mark_status(self, row_index, status_color, date_column, status_name):
        """
        Internal helper method to mark status with a specific color and update date column.
        Synchronously updates both cached workbook and _internal_status in DataFrame.

        Args:
            row_index: The row index to mark (1-indexed, including header row)
            status_color: RGB color code (e.g., 'FFFF0000' for red)
            date_column: Name of the date column to update ('Processed Date' or 'Result Date')
            status_name: Name of the status for logging (e.g., 'rejected', 'processing', 'offer')

        Returns:
            True if successful, False otherwise
        """
        if row_index is None or row_index < 2:  # Row 1 is header
            print_("Invalid row index.", "RED")
            return False

        try:
            # Check for write conflict before marking
            if not self._check_for_write_conflict():
                return False

            # Create backup using system temp directory
            temp_dir = tempfile.gettempdir()
            backup_filename = os.path.join(temp_dir, f"job_application_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            shutil.copy2(self.file_path, backup_filename)
            print_(f"\nBackup created at {backup_filename}")

            # Use cached workbook or load new one
            if self._cached_workbook is None:
                workbook = load_workbook(filename=self.file_path)
                self._cached_workbook = workbook
            else:
                workbook = self._cached_workbook

            sheet = workbook.active

            # Find the Status and date columns
            headers = {cell.value: cell.column for cell in sheet[1]}

            if 'Status' not in headers:
                raise ValueError("[DEBUG] 'Status' column not found in the Excel file.")

            # Mark the Status cell with specified color
            status_cell = sheet.cell(row=row_index, column=headers['Status'])
            fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            status_cell.fill = fill

            # Update date column
            current_date = datetime.now().strftime('%Y-%m-%d')
            if date_column in headers:
                date_cell = sheet.cell(row=row_index, column=headers[date_column])
                date_cell.value = current_date
            else:
                raise ValueError(f"[DEBUG] {date_column} column not found in the Excel file.")

            # Update cached DataFrame to keep it in sync
            df_index = row_index - 2  # Convert Excel row to DataFrame index
            if self._cached_df is not None and df_index < len(self._cached_df):
                internal_status = self._color_to_status(status_color)
                self._cached_df.at[df_index, '_internal_status'] = internal_status

                # Update date column in DataFrame as well
                if date_column in self._cached_df.columns:
                    self._cached_df.at[df_index, date_column] = current_date

            print_(f"Row {row_index - 1} marked as {status_name}.", "GREEN")
            return True

        except Exception as e:
            print_(f"Error marking as {status_name}: {str(e)}", "RED")
            return False

    def mark_as_rejected(self, row_index=None):
        """
        Marks the 'Status' cell at the given row index with a red fill color.
        Updates the 'Result Date' column with current date.
        Also creates a backup of the file in system temp directory.

        Args:
            row_index: The row index to mark (1-indexed, including header row)

        Returns:
            True if successful, False otherwise
        """
        return self._mark_status(row_index, self.STATUS_COLORS['REJECTED'], "Result Date", "REJECTED")

    def mark_as_processing(self, row_index=None):
        """
        Marks the 'Status' cell at the given row index with a yellow fill color.
        Updates the 'Processed Date' column with current date.
        Also creates a backup of the file in system temp directory.

        Args:
            row_index: The row index to mark (1-indexed, including header row)

        Returns:
            True if successful, False otherwise
        """
        return self._mark_status(row_index, self.STATUS_COLORS['PROCESSING'], "Processed Date", "PROCESSING")

    def mark_as_offer(self, row_index=None):
        """
        Marks the 'Status' cell at the given row index with a green fill color.
        Updates the 'Result Date' column with current date.
        Also creates a backup of the file in system temp directory.

        Args:
            row_index: The row index to mark (1-indexed, including header row)

        Returns:
            True if successful, False otherwise
        """
        return self._mark_status(row_index, self.STATUS_COLORS['OFFER'], "Result Date", "OFFER")
