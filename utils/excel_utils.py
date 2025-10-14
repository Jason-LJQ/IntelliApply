import re

from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import shutil
from datetime import datetime

from utils.print_utils import print_, print_results
from utils.string_utils import normalize_company_name, get_abbreviation_lower, cleaned_string
from config.credential import EXCEL_FILE_PATH
from config.prompt import ALL_FIELDS


def validate_excel_file(excel_file=EXCEL_FILE_PATH):
    """
    Validates the Excel file:
    1. Check if file exists, if not, ask user to create
    2. Check if all required columns exist, if not, backup the file and create a new one
    
    Returns:
        bool: True if validation passes or issues are fixed, False if validation fails
    """

    def create_new_excel():
        wb = Workbook()
        ws = wb.active
        for col, field in enumerate(['Result'] + ALL_FIELDS, 1):
            ws.cell(row=1, column=col, value=field)
        wb.save(excel_file)
        print_(f"Created new Excel file at {excel_file}", "GREEN")

    try:
        # Check if file exists
        if not os.path.exists(excel_file):
            dir_path = os.path.dirname(excel_file)
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)

            confirm = input(
                print_(f"Excel file not found at {excel_file}. Create it? (y/Y to confirm): ", color="YELLOW",
                       return_text=True)).lower()
            if confirm == 'y':
                # Create new Excel file with required columns
                create_new_excel()
                return True
            else:
                print_("Excel file creation cancelled.", "RED")
                return False

        # Check if all required columns exist
        wb = load_workbook(filename=excel_file)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        missing_headers = [header for header in ['Result'] + ALL_FIELDS if header not in existing_headers]
        wb.close()

        if missing_headers:
            # Confirm to backup and create new file
            confirm = input(
                print_(f"Missing columns: {missing_headers}. Backup and create new file? (y/Y to confirm): ",
                       color="YELLOW", return_text=True)).lower()
            if confirm != 'y':
                print_("Validation failed.", "RED")
                return False

            # Backup existing file using shutil
            backup_path = f"{os.path.splitext(excel_file)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_bak.xlsx"
            shutil.move(excel_file, backup_path)
            print_(f"Backed up existing file to {backup_path}", "YELLOW")

            # Create new file with correct headers
            create_new_excel()

            return True

        return True

    except Exception as e:
        print_(f"Error validating Excel file: {str(e)}", "RED")
        return False


def get_result_status(workbook, row_index):
    """
    Checks if the 'Result' cell at the given row index has any fill color.
    Returns 'x' if any fill color is found, otherwise returns an empty string.
    """
    try:
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}

        if 'Result' not in headers:
            return ''

        result_cell = sheet.cell(row=row_index, column=headers['Result'])

        # Check if the cell has any fill color
        if result_cell.fill.start_color.rgb == '00000000':
            return ''
        else:
            return '  ⨉  '

    except Exception as e:
        print_(f"Error reading cell color: {str(e)}", "RED")
        return ''


def show_last_row(excel_file=EXCEL_FILE_PATH, delete=False):
    """
    Show last row's data. If delete is True, delete the last row after user confirmation.
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        last_row = sheet.max_row

        if last_row <= 1:
            print_("No data to show.", "RED")
            workbook.close()
            return

        # Use search_applications to get the last row data
        results = search_applications(excel_file=excel_file, index=last_row)
        print_results(results)

        if delete:
            # Ask for user confirmation
            confirm = input(print_("Delete this row? (y/Y to confirm, any other key to cancel): ", color="YELLOW",
                                   return_text=True)).lower()
            if confirm == 'y':
                sheet.delete_rows(last_row)
                workbook.save(filename=excel_file)
                print_(f"Last row deleted successfully.", "GREEN")
            else:
                print_(f"Deletion cancelled.", "RED")

        workbook.close()

    except Exception as e:
        print_(f"Error processing last row: \n{str(e)}", "RED")


def search_applications(excel_file=EXCEL_FILE_PATH, search_term="", index=-1):
    """
    Search for both company and job title matches using vectorized pandas operations,
    ensuring full compatibility with the original matching logic.
    
    Args:
        excel_file: Path to the Excel file
        search_term: The search term to match against company and job title
        index: If >= 0, directly return the record at this index (1-indexed, including header row)
        
    Returns:
        List of matching records
    """
    try:
        # Read Excel file fresh every time
        df = pd.read_excel(excel_file)
        workbook = load_workbook(filename=excel_file)

        def applied_date(row):
            raw = row.get('Applied Date', '')
            # Handle pandas NaT (Not a Time) or numpy.nan
            return '' if pd.isna(raw) else raw

        # If index is provided, directly return that record
        if index >= 2:  # Index 1 is header
            if index - 2 < len(df):  # Convert to 0-indexed for DataFrame
                row = df.iloc[index - 2]

                result = [{
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row),
                    'result': get_result_status(workbook, index),
                    'row_index': index,
                }]
                workbook.close()
                return result
            else:
                print_(f"Index {index} is out of range.", "RED")
                return []

        # 1. Prepare search term variations
        search_term_clean_lower = cleaned_string(search_term).lower()
        norm_keyword = normalize_company_name(search_term)
        abbr_keyword = get_abbreviation_lower(norm_keyword)

        # Ensure target columns are string type
        df['Company'] = df['Company'].astype(str).fillna('')
        df['Job Title'] = df['Job Title'].astype(str).fillna('')

        # 2. Create helper columns using vectorized `apply`
        df['norm_company'] = df['Company'].apply(normalize_company_name)
        df['abbr_company'] = df['Company'].apply(get_abbreviation_lower)
        df['clean_job_title'] = df['Job Title'].apply(cleaned_string).str.lower()

        # Create a regex pattern for word-level matching
        job_title_pattern = r'\b' + re.escape(search_term_clean_lower) + r'\b'

        # 3. Build boolean masks that perfectly replicate the original function's logic

        # Mask 1: Direct, Prefix, and Job Title matches
        m_base = (
                (df['norm_company'] == norm_keyword) |
                (df['norm_company'].str.startswith(norm_keyword, na=False)) |
                (df['clean_job_title'].str.contains(job_title_pattern, na=False, regex=True))
        )

        # Mask 2: Handles "om" matching "Old Mission" (Abbreviation of DB entry matches the full keyword)
        # This is the key logic that was previously broken. It is NOT gated by length.
        m_abbr_target = (df['abbr_company'] == norm_keyword)

        # Mask 3: Handles "GSK" matching "GlaxoSmithKline" (Abbreviation of keyword matches abbreviation of DB entry)
        # This part IS gated by length to prevent single-letter false positives like 'd' matching 'Databricks'.
        m_abbr_keyword_vs_abbr_target = pd.Series([False] * len(df), index=df.index)
        if len(abbr_keyword) > 1:
            m_abbr_keyword_vs_abbr_target = (df['abbr_company'] == abbr_keyword)

        # 4. Combine all masks using logical OR
        final_mask = m_base | m_abbr_target | m_abbr_keyword_vs_abbr_target

        # 5. Filter the DataFrame and format results
        matched_df = df[final_mask]

        if matched_df.empty:
            workbook.close()
            return []

        matches = []
        for idx, row in matched_df.iterrows():
            excel_row_index = idx + 2  # Convert 0-based DataFrame index to 1-based Excel row index (with header)
            matches.append({
                'Company': row['Company'],
                'Location': row['Location'],
                'Job Title': row['Job Title'],
                'Applied Date': applied_date(row.to_dict()),
                'result': get_result_status(workbook, excel_row_index),
                'row_index': excel_row_index,
            })

        workbook.close()
        return matches

    except Exception as e:
        print_(f"Error during search: {str(e)}", "RED")
        return []


def append_data_to_excel(excel_file=EXCEL_FILE_PATH, data=None):
    """
    Appends a list of dictionaries to the end of the Excel file using openpyxl.
    Automatically determines the column letters for all fields.
    Adds missing columns if they don't exist.
    """
    if data is None:
        return
    try:
        # Load the existing Excel file
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Find the last row with data (or the header row if the sheet is empty)
        last_row = sheet.max_row if sheet.max_row > 1 else 1

        # Get the headers and their column indices
        headers = {cell.value: cell.column for cell in sheet[1]}

        # List of all possible columns
        all_columns = ALL_FIELDS

        # Add any missing columns
        next_column = len(headers) + 1
        for column in all_columns:
            if column not in headers:
                headers[column] = next_column
                sheet.cell(row=1, column=next_column, value=column)
                next_column += 1

        # Append the new data directly to the last row
        for row_data in data:
            last_row += 1  # Move to the next row for each new record

            # Write data to the corresponding columns
            for column in all_columns:
                if column != 'Result':  # Skip Result column as it's handled separately
                    sheet.cell(row=last_row, column=headers[column],
                               value=row_data.get(column, ''))

        # Save the updated workbook
        workbook.save(filename=excel_file)

    except Exception as e:
        print_(f"Error appending data to Excel: {str(e)}")


def check_duplicate_entry(excel_file=EXCEL_FILE_PATH, new_data=None):
    """
    Check if the exact same job entry already exists in the Excel file.
    Returns True if a duplicate is found, False otherwise.
    """
    if new_data is None:
        return None
    try:
        df = pd.read_excel(excel_file)

        # Convert all columns to string for comparison
        for col in df.columns:
            df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        # Check each row for exact match
        for _, row in df.iterrows():
            all_fields_match = True
            for field in ['Company', 'Job Title']:
                if field in row and field in new_data:
                    if str(row[field]).strip() != str(new_data[field]).strip():
                        all_fields_match = False
                        break
            if all_fields_match:
                return row
        return None

    except Exception as e:
        print_(f"Error checking for duplicates: {str(e)}", "RED")
        return None


def summary(excel_file=EXCEL_FILE_PATH):
    """
    Print a summary of job applications including:
    - Total number of applications
    - Number of rejections
    - Rejection percentage
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Get total number of applications (excluding header row)
        total_applications = sheet.max_row - 1

        # Count rejections by checking cell fill color
        rejections = 0
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
            result = get_result_status(workbook, row)
            if result.strip():  # If result has an 'x' mark
                rejections += 1

        # Calculate rejection percentage
        rejection_percentage = (rejections / total_applications * 100) if total_applications > 0 else 0

        # Print summary with color formatting
        print_(f"\nApplication Summary:")
        print(f"Total Applications: {total_applications}")
        print(f"Rejections: {rejections}")
        print(f"Rejection Rate: {rejection_percentage:.1f}%")

        workbook.close()

    except Exception as e:
        print_(f"Error generating summary: {str(e)}", "RED")


def open_excel_file(excel_file=EXCEL_FILE_PATH):
    """
    Opens the Excel file using the default application based on the operating system.
    Returns True if successful, False otherwise.
    """
    try:
        import subprocess
        import platform
        system = platform.system()
        if system == 'Darwin':  # macOS
            subprocess.run(['open', excel_file])
        elif system == 'Windows':
            subprocess.run(['start', '', excel_file], shell=True)
        else:  # Linux and other OS
            subprocess.run(['xdg-open', excel_file])
        print_(f"Excel file opened successfully.", "GREEN")
    except Exception as e:
        print_(f"Error opening Excel file: {str(e)}", "RED")
        return False
    return True


def mark_result(excel_file=EXCEL_FILE_PATH, row_index=None):
    """
    Marks the 'Result' cell at the given row index with a red fill color.
    Also creates a backup of the file in /tmp directory.
    
    Args:
        excel_file: Path to the Excel file
        row_index: The row index to mark (1-indexed, including header row)
        
    Returns:
        True if successful, False otherwise
    """
    import shutil
    import os
    from datetime import datetime
    from openpyxl.styles import PatternFill

    if row_index is None or row_index < 2:  # Row 1 is header
        print_("Invalid row index.", "RED")
        return False

    try:
        # Create backup in /tmp directory
        backup_filename = f"/tmp/job_application_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(excel_file, backup_filename)
        print_(f"\nBackup created at {backup_filename}")

        # Load the workbook
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Find the Result column
        headers = {cell.value: cell.column for cell in sheet[1]}

        if 'Result' not in headers:
            print_("'Result' column not found in the Excel file.", "RED")
            workbook.close()
            return False

        # Mark the cell with red fill
        result_cell = sheet.cell(row=row_index, column=headers['Result'])
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        result_cell.fill = red_fill

        # Save the workbook
        workbook.save(filename=excel_file)
        workbook.close()

        print_(f"Row {row_index - 1} marked as rejected.", "GREEN")
        return True

    except Exception as e:
        print_(f"Error marking result: {str(e)}", "RED")
        return False
