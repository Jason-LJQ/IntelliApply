from openpyxl import load_workbook
import pandas as pd
from print_utils import print_
from string_utils import is_company_match

from config import EXCEL_FILE_PATH


def get_result_status(workbook, row_index):
    """
    Checks if the 'Result' cell at the given row index has any fill color.
    Returns 'x' if any fill color is found, otherwise returns an empty string.
    """
    try:
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}

        if 'Result' not in headers:
            return ''

        result_cell = sheet.cell(row=row_index, column=headers['Result'])

        # Check if the cell has any fill color
        if result_cell.fill.start_color.rgb == '00000000':
            return ''
        else:
            return '  ⨉  '

    except Exception as e:
        print_(f"Error reading cell color: {str(e)}", "RED")
        return ''


def delete_last_row(excel_file=EXCEL_FILE_PATH):
    """
    Deletes the last row in the Excel file after user confirmation.
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        last_row = sheet.max_row

        if last_row <= 1:
            print_("No data to delete.", "RED")
            workbook.close()
            return

        # Display the last row's data for confirmation
        last_row_data = [cell.value for cell in sheet[last_row]]
        headers = [cell.value for cell in sheet[1]]
        print_("Last row data:")
        for header, value in zip(headers, last_row_data):
            print(f"{header}: {value}")

        # Ask for user confirmation
        confirm = input("[*] Delete this row? (y/Y to confirm, any other key to cancel): ").lower()
        if confirm == 'y':
            sheet.delete_rows(last_row)
            workbook.save(filename=excel_file)
            print_(f"Last row deleted successfully.", "GREEN")
        else:
            print_(f"Deletion cancelled.", "RED")

        workbook.close()

    except Exception as e:
        print_(f"Error deleting last row: {str(e)}", "RED")


def search_applications(excel_file=EXCEL_FILE_PATH, search_term=""):
    """
    Search for both company and job title matches
    """
    try:
        # Read Excel file fresh every time
        df = pd.read_excel(excel_file)

        # Convert DataFrame columns to string type
        string_columns = ['Company', 'Job Title', 'Location']
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        matches = []
        search_term_lower = search_term.lower().strip()
        workbook = load_workbook(filename=excel_file)

        def applied_date(row):
            raw = row.get('Applied Date', '')
            raw = '' if str(raw).strip() == 'nan' else raw
            return raw

        for index, row in df.iterrows():
            # Check company name match
            if is_company_match(search_term, row['Company']):
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row),
                    'result': get_result_status(workbook, index + 2),
                })
            # Check exact job title match
            elif search_term_lower in row['Job Title'].lower():
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row),
                    'result': get_result_status(workbook, index + 2),
                })

        workbook.close()
        return matches

    except Exception as e:
        print_(f"Error reading file: {str(e)}")
        return []


def append_data_to_excel(excel_file=EXCEL_FILE_PATH, data=[]):
    """
    Appends a list of dictionaries to the end of the Excel file using openpyxl.
    Automatically determines the column letters for all fields.
    Adds missing columns if they don't exist.
    """
    try:
        # Load the existing Excel file
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Find the last row with data (or the header row if the sheet is empty)
        last_row = sheet.max_row if sheet.max_row > 1 else 1

        # Get the headers and their column indices
        headers = {cell.value: cell.column for cell in sheet[1]}

        # List of all possible columns
        from config import ALL_FIELDS
        all_columns = ALL_FIELDS

        # Add any missing columns
        next_column = len(headers) + 1
        for column in all_columns:
            if column not in headers:
                headers[column] = next_column
                sheet.cell(row=1, column=next_column, value=column)
                next_column += 1

        # Append the new data directly to the last row
        for row_data in data:
            last_row += 1  # Move to the next row for each new record

            # Write data to the corresponding columns
            for column in all_columns:
                if column != 'Result':  # Skip Result column as it's handled separately
                    sheet.cell(row=last_row, column=headers[column],
                               value=row_data.get(column, ''))

        # Save the updated workbook
        workbook.save(filename=excel_file)

    except Exception as e:
        print_(f"Error appending data to Excel: {str(e)}")


def check_duplicate_entry(excel_file, new_data):
    """
    Check if the exact same job entry already exists in the Excel file.
    Returns True if a duplicate is found, False otherwise.
    """
    try:
        df = pd.read_excel(excel_file)

        # Convert all columns to string for comparison
        for col in df.columns:
            df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        # Check each row for exact match
        for _, row in df.iterrows():
            all_fields_match = True
            for field in ['Company', 'Job Title']:
                if field in row and field in new_data:
                    if str(row[field]).strip() != str(new_data[field]).strip():
                        all_fields_match = False
                        break
            if all_fields_match:
                return row
        return None

    except Exception as e:
        print_(f"Error checking for duplicates: {str(e)}", "RED")
        return None


def summary(excel_file=EXCEL_FILE_PATH):
    """
    Print a summary of job applications including:
    - Total number of applications
    - Number of rejections
    - Rejection percentage
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Get total number of applications (excluding header row)
        total_applications = sheet.max_row - 1

        # Count rejections by checking cell fill color
        rejections = 0
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
            result = get_result_status(workbook, row)
            if result.strip():  # If result has an 'x' mark
                rejections += 1

        # Calculate rejection percentage
        rejection_percentage = (rejections / total_applications * 100) if total_applications > 0 else 0

        # Print summary with color formatting
        print_(f"\nApplication Summary:")
        print(f"Total Applications: {total_applications}")
        print(f"Rejections: {rejections}")
        print(f"Rejection Rate: {rejection_percentage:.1f}%")

        workbook.close()

    except Exception as e:
        print_(f"Error generating summary: {str(e)}", "RED")


def open_excel_file(excel_file=EXCEL_FILE_PATH):
    """
    Opens the Excel file using the default application based on the operating system.
    Returns True if successful, False otherwise.
    """
    try:
        import subprocess
        import platform
        system = platform.system()
        if system == 'Darwin':  # macOS
            subprocess.run(['open', excel_file])
        elif system == 'Windows':
            subprocess.run(['start', '', excel_file], shell=True)
        else:  # Linux and other OS
            subprocess.run(['xdg-open', excel_file])
        print_("Opening Excel file...", "GREEN")
        return True
    except Exception as e:
        print_(f"Error opening file: {str(e)}", "RED")
        return False
