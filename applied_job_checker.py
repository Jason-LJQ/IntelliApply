############################################
# Author: Jason Liao
# Date: 2024-12-22
# Description: A simple script to search for job applications in an Excel file
############################################

import pandas as pd
import re
import shutil
import signal
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_abbreviation(name):
    """Get the abbreviation of a string by taking first letters of each word"""
    if not isinstance(name, str):
        return ''
    # Remove special characters, keep only letters and spaces
    cleaned_name = re.sub(r'[^a-zA-Z\s]', '', str(name))
    words = cleaned_name.strip().split()
    abbr_parts = []
    for w in words:
        if w.isupper():
            abbr_parts.append(w)  # Preserve the entire uppercase word
        else:
            abbr_parts.append(w[0].upper() if w else '')
    return ''.join(abbr_parts)


def normalize_company_name(name):
    """Normalize company name by removing common suffixes and extra spaces"""
    if not isinstance(name, str):
        return ''
    # Remove common company terms
    common_terms = ['corporation', 'corp', 'inc', 'incorporated', 'limited', 'ltd', 'llc', 'cooperation']
    name_lower = str(name).lower().strip()
    for term in common_terms:
        name_lower = re.sub(rf'\b{term}\b', '', name_lower)
    # Remove special characters and extra spaces
    name_lower = re.sub(r'[^a-zA-Z\s]', '', name_lower)
    return re.sub(r'\s+', ' ', name_lower).strip()


def format_location(location):
    """Convert multi-line location to single line with semicolons"""
    formatted = '; '.join(str(location).strip().split('\n'))
    # Limit location length
    if len(formatted) > 70:
        return formatted[:70] + '...'
    return formatted


def get_result_status(workbook, row_index):
    """
    Checks if the 'Result' cell at the given row index has any fill color.
    Returns 'x' if any fill color is found, otherwise returns an empty string.
    """
    try:
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}

        if 'Result' not in headers:
            return ''

        result_cell = sheet.cell(row=row_index, column=headers['Result'])

        # Check if the cell has any fill color
        if result_cell.fill.start_color.rgb == '00000000':
            return ''
        else:
            return '  ⨉  '

    except Exception as e:
        print(f"Error reading cell color: {str(e)}")
        return ''


def print_results(results, excel_file):
    """
    Prints the search results, including the 'Result' column with 'x' for colored cells.
    """
    if not results:
        print("\nNot found.")
        return

    print(f"\nFound {len(results)} matching records:")

    # Calculate maximum widths for each column, including 'Result'
    company_width = max(len(str(r['Company'])) for r in results)
    company_width = max(company_width, len("Company"))

    location_width = max(len(format_location(r['Location'])) for r in results)
    location_width = max(location_width, len("Location"))

    job_width = max(len(str(r['Job Title'])) for r in results)
    job_width = max(job_width, len("Job Title"))

    result_width = max(len(str(r.get('Result', ''))) for r in results)
    result_width = max(result_width, len("Result"))

    # Print header, including 'Result'
    print("\n{:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
        "Result", "Company", "Location", "Job Title",
        width1=result_width,
        width2=company_width,
        width3=location_width,
        width4=job_width
    ))
    print("-" * (company_width + location_width + job_width + result_width + 4))

    # Print results
    for result in results:
        print("{:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
            str(result['result']),
            str(result['Company']),
            format_location(result['Location']),
            str(result['Job Title']),
            width1=result_width,
            width2=company_width,
            width3=location_width,
            width4=job_width
        ))


def is_company_match(keyword, target):
    """
    Compare if two company names match, considering exact matches and abbreviations
    """
    if not isinstance(keyword, str) or not isinstance(target, str):
        return False

    # Normalize both company names
    norm_keyword = normalize_company_name(keyword)
    norm_target = normalize_company_name(target)

    # Direct comparison after normalization
    if norm_keyword == norm_target:
        return True

    # Get abbreviations
    abbr1 = get_abbreviation(keyword)
    abbr2 = get_abbreviation(target)

    # Compare abbreviation with full name and vice versa
    if len(abbr1) > 1:
        if abbr1 == abbr2:
            return True
        # Check if abbr1 matches the first letters of norm_target's words
        if abbr1 == get_abbreviation(norm_target):
            return True

    if len(abbr2) > 1:
        # Check if abbr2 matches the first letters of norm_keyword's words
        if abbr2 == get_abbreviation(norm_keyword):
            return True

    # Additional comparison: check if abbreviation of one matches normalized other
    abbr1 = get_abbreviation(norm_keyword)
    abbr2 = get_abbreviation(norm_target)

    if abbr1 == norm_target or abbr2 == norm_keyword:
        return True

    # Handle case where one is abbreviation and other is full name
    words1 = norm_keyword.split()
    words2 = norm_target.split()

    if len(words1) >= 2 and len(words2) >= 2:
        # Check if the initials of one match the other's abbreviation
        initials1 = ''.join(word[0].upper() for word in words1)
        initials2 = ''.join(word[0].upper() for word in words2)

        if initials1 == abbr2 or initials2 == abbr1:
            return True

    # Check if keyword is the substring of target from the beginning
    if norm_keyword and norm_target:
        if norm_target.startswith(norm_keyword):
            return True

    return False


def search_applications(excel_file, search_term):
    """
    Search for both company and job title matches
    """
    try:
        # Read Excel file fresh every time
        df = pd.read_excel(excel_file)

        # Convert DataFrame columns to string type
        string_columns = ['Company', 'Job Title', 'Location']
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        matches = []
        search_term_lower = search_term.lower().strip()
        workbook = load_workbook(filename=excel_file)

        for index, row in df.iterrows():
            # Check company name match
            if is_company_match(search_term, row['Company']):
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'result': get_result_status(workbook, index + 2),
                })
            # Check exact job title match
            elif search_term_lower in row['Job Title'].lower():
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'result': get_result_status(workbook, index + 2),
                })

        workbook.close()
        return matches

    except Exception as e:
        print(f"Error reading file: {str(e)}")
        return []


def parse_markdown_table(markdown_table_string):
    """
    Parses a Markdown table string and returns a list of dictionaries,
    where each dictionary represents a row in the table.
    """
    lines = markdown_table_string.strip().split('\n')
    if len(lines) < 3:
        return []  # Not a valid Markdown table

    # Extract headers
    headers = [header.strip() for header in lines[0].split('|') if header.strip()]

    # Extract rows
    data = []
    for line in lines[2:]:
        values = [value.strip() for value in line.split('|') if value.strip()]
        if len(values) == len(headers):
            data.append(dict(zip(headers, values)))
    return data


def append_data_to_excel(excel_file, data):
    """
    Appends a list of dictionaries to the end of the Excel file using openpyxl.
    Automatically determines the column letters for "Company", "Location", "Job Title", and "Result".
    Adds a blank "Result" column if it doesn't exist.
    """
    try:
        # Load the existing Excel file
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Find the last row with data (or the header row if the sheet is empty)
        last_row = sheet.max_row if sheet.max_row > 1 else 1

        # Get the headers and their column indices
        headers = {cell.value: cell.column for cell in sheet[1]}

        # Add a blank "Result" column if it doesn't exist
        if 'Result' not in headers:
            headers['Result'] = len(headers) + 1
            sheet.cell(row=1, column=headers['Result'], value='Result')

        # Append the new data directly to the last row
        for row_data in data:
            last_row += 1  # Move to the next row for each new record

            # Write data to the corresponding columns
            sheet.cell(row=last_row, column=headers.get('Company', 1), value=row_data.get('Company', ''))
            sheet.cell(row=last_row, column=headers.get('Location', 1), value=row_data.get('Location', ''))
            sheet.cell(row=last_row, column=headers.get('Job Title', 1), value=row_data.get('Job Title', ''))

        # Save the updated workbook
        workbook.save(filename=excel_file)

    except Exception as e:
        print(f"Error appending data to Excel: {str(e)}")


def signal_handler(sig, frame):
    print('\nExiting ...')
    sys.exit(0)


def delete_last_row(excel_file):
    """
    Deletes the last row in the Excel file after user confirmation.
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        last_row = sheet.max_row

        if last_row <= 1:
            print("No data to delete.")
            workbook.close()
            return

        # Display the last row's data for confirmation
        last_row_data = [cell.value for cell in sheet[last_row]]
        headers = [cell.value for cell in sheet[1]]
        print("Last row data:")
        for header, value in zip(headers, last_row_data):
            print(f"{header}: {value}")

        # Ask for user confirmation
        confirm = input("Delete this row? (y/Y to confirm, any other key to cancel): ").lower()
        if confirm == 'y':
            sheet.delete_rows(last_row)
            workbook.save(filename=excel_file)
            print("Last row deleted successfully.")
        else:
            print("Deletion cancelled.")

        workbook.close()

    except Exception as e:
        print(f"Error deleting last row: {str(e)}")


def main(excel_file):
    # Set up signal handler for SIGINT
    signal.signal(signal.SIGINT, signal_handler)

    try:
        while True:
            print("\nEnter search keyword, paste Markdown table, 'delete' to delete last row (or 'exit' to quit):")
            user_input_lines = []
            line_count = 0
            while line_count < 3:
                line = input()
                if not line:
                    break
                if "|" not in line:  # Not a Markdown table
                    user_input_lines.append(line)
                    break
                user_input_lines.append(line)
                line_count += 1

            user_input = '\n'.join(user_input_lines).strip()

            if user_input.lower() == 'exit':
                print("Exiting ...")
                break

            if user_input.lower() == 'delete':
                delete_last_row(excel_file)
                continue
            
            if not user_input:
                print("Search keyword cannot be empty!")
                continue

            if is_markdown_table(user_input):
                # Parse the Markdown table
                data = parse_markdown_table(user_input)

                if not data:
                    print("Invalid Markdown table format.")
                    continue

                # Append the data to the Excel file
                append_data_to_excel(excel_file, data)
                print("New record successfully appended to Excel file.")

            else:
                results = search_applications(excel_file, user_input)
                if results:
                    print_results(results, excel_file)
                else:
                    print("\nNo matching records found.")

    except KeyboardInterrupt:
        print('\nExiting ...')
        sys.exit(0)


def is_markdown_table(input_string):
    """
    Checks if the input string is likely a Markdown table.
    This is a heuristic check and may not be 100% accurate.
    """
    lines = input_string.strip().split('\n')
    if len(lines) < 3:
        return False  # Not enough lines for a table

    # Check for at least one | in the header and data lines
    if '|' not in lines[0]:
        return False

    for line in lines[2:]:
        if '|' not in line:
            return False

    return True


if __name__ == "__main__":
    main(
        excel_file='/Users/jason/Library/CloudStorage/OneDrive-Personal/Graduate Study/17-677-I Internship for Software Engineers - Summer 2025/Book1.xlsx')
