############################################
# Author: Jason Liao
# Date: 2024-12-22
# Description: A simple script to search for job applications in an Excel file
############################################
import os
import subprocess
import time
import pandas as pd
import re
import shutil
import signal
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openai import OpenAI
from config import *
from credential import *
import json
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse  # Add color constants at the top after imports

# Color constants
RED = '\033[31m'
GREEN = '\033[32m'
RESET = '\033[0m'

COOKIE_PATH = os.path.join(os.path.dirname(__file__), "cookie.txt")
DOMAIN_KEYWORDS = {"https://www.linkedin.com/mypreferences/d/categories/account": ["preferred", "demographic"],
                   "https://app.joinhandshake.com": ["explore", "people"]}

client = OpenAI(api_key=OPENAI_API_KEY, base_url=BASE_URL)


def get_abbreviation(name):
    """Get the abbreviation of a string by taking first letters of each word"""
    if not isinstance(name, str):
        return ''
    # Remove special characters, keep only letters and spaces
    cleaned_name = re.sub(r'[^a-zA-Z\s]', '', str(name))
    words = cleaned_name.strip().split()
    abbr_parts = []
    for w in words:
        if w.isupper():
            abbr_parts.append(w)  # Preserve the entire uppercase word
        else:
            abbr_parts.append(w[0].upper() if w else '')
    return ''.join(abbr_parts)


def normalize_company_name(name):
    """Normalize company name by removing common suffixes and extra spaces"""
    if not isinstance(name, str):
        return ''
    # Remove common company terms
    common_terms = ['corporation', 'corp', 'inc', 'incorporated', 'limited', 'ltd', 'llc', 'cooperation']
    name_lower = str(name).lower().strip()
    for term in common_terms:
        name_lower = re.sub(rf'\b{term}\b', '', name_lower)
    # Remove special characters and extra spaces
    name_lower = re.sub(r'[^a-zA-Z\s]', '', name_lower)
    return re.sub(r'\s+', ' ', name_lower).strip()


def format_string(name, limit=55):
    """Convert multi-line location to single line with semicolons"""
    formatted = '; '.join(str(name).strip().split('\n'))
    # Limit location length
    if len(formatted) > limit:
        return formatted[:limit] + '...'
    return formatted


def get_result_status(workbook, row_index):
    """
    Checks if the 'Result' cell at the given row index has any fill color.
    Returns 'x' if any fill color is found, otherwise returns an empty string.
    """
    try:
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}

        if 'Result' not in headers:
            return ''

        result_cell = sheet.cell(row=row_index, column=headers['Result'])

        # Check if the cell has any fill color
        if result_cell.fill.start_color.rgb == '00000000':
            return ''
        else:
            return '  ⨉  '

    except Exception as e:
        print(f"Error reading cell color: {str(e)}")
        return ''


def print_results(results, excel_file):
    """
    Prints the search results, including the 'Result' column with 'x' for colored cells.
    """
    if not results:
        print(f"\n{RED}[*] Not found.{RESET}")
        return

    print(f"\n{GREEN}[*] Found {len(results)} matching records:{RESET}")

    # Calculate maximum widths for each column, including 'Result'
    company_width = max(len(format_string(r['Company'], limit=30)) for r in results)
    company_width = max(company_width, len("Company"))

    location_width = max(len(format_string(r['Location'])) for r in results)
    location_width = max(location_width, len("Location"))

    job_width = max(len(format_string(r['Job Title'], limit=65)) for r in results)
    job_width = max(job_width, len("Job Title"))

    result_width = max(len(str(r.get('Result', ''))) for r in results)
    result_width = max(result_width, len("Result"))

    # Check if any Applied Date is valid
    has_valid_date = any(str(r.get('Applied Date', '')).strip() != '' for r in results)
    if has_valid_date:
        date_width = max(len(str(r.get('Applied Date', ''))) for r in results)
        date_width = max(date_width, len("Applied Date"))

    # Print header, including 'Result' and 'Applied Date' if valid
    if has_valid_date:
        print("\n{:<{width5}}  {:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
            "Applied Date", "Result", "Company", "Location", "Job Title",
            width5=date_width,
            width1=result_width,
            width2=company_width,
            width3=location_width,
            width4=job_width
        ))
        print("-" * (company_width + location_width + job_width + result_width + date_width + 8))
    else:
        print("\n{:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
            "Result", "Company", "Location", "Job Title",
            width1=result_width,
            width2=company_width,
            width3=location_width,
            width4=job_width
        ))
        print("-" * (company_width + location_width + job_width + result_width + 6))

    # Print results
    for result in results:
        if has_valid_date:
            print("{:<{width5}}  {:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
                str(result.get('Applied Date', '')),
                str(result['result']),
                format_string(result['Company'], limit=30),
                format_string(result['Location']),
                format_string(result['Job Title'], limit=65),
                width5=date_width,
                width1=result_width,
                width2=company_width,
                width3=location_width,
                width4=job_width
            ))
        else:
            print("{:<{width1}}  {:<{width2}}  {:<{width3}}  {:<{width4}}".format(
                str(result['result']),
                format_string(result['Company'], limit=30),
                format_string(result['Location']),
                format_string(result['Job Title'], limit=65),
                width1=result_width,
                width2=company_width,
                width3=location_width,
                width4=job_width
            ))


def is_company_match(keyword, target):
    """
    Compare if two company names match, considering exact matches and abbreviations
    """
    if not isinstance(keyword, str) or not isinstance(target, str):
        return False

    # Normalize both company names
    norm_keyword = normalize_company_name(keyword)
    norm_target = normalize_company_name(target)

    # Direct comparison after normalization
    if norm_keyword == norm_target:
        return True

    # Get abbreviations
    abbr1 = get_abbreviation(keyword)
    abbr2 = get_abbreviation(target)

    # Compare abbreviation with full name and vice versa
    if len(abbr1) > 1:
        if abbr1 == abbr2:
            return True
        # Check if abbr1 matches the first letters of norm_target's words
        if abbr1 == get_abbreviation(norm_target):
            return True

    if len(abbr2) > 1:
        # Check if abbr2 matches the first letters of norm_keyword's words
        if abbr2 == get_abbreviation(norm_keyword):
            return True

    # Additional comparison: check if abbreviation of one matches normalized other
    abbr1 = get_abbreviation(norm_keyword)
    abbr2 = get_abbreviation(norm_target)

    if abbr1 == norm_target or abbr2 == norm_keyword:
        return True

    # Handle case where one is abbreviation and other is full name
    words1 = norm_keyword.split()
    words2 = norm_target.split()

    if len(words1) >= 2 and len(words2) >= 2:
        # Check if the initials of one match the other's abbreviation
        initials1 = ''.join(word[0].upper() for word in words1)
        initials2 = ''.join(word[0].upper() for word in words2)

        if initials1 == abbr2 or initials2 == abbr1:
            return True

    # Check if keyword is the substring of target from the beginning
    if norm_keyword and norm_target:
        if norm_target.startswith(norm_keyword):
            return True

    return False


def search_applications(excel_file, search_term):
    """
    Search for both company and job title matches
    """
    try:
        # Read Excel file fresh every time
        df = pd.read_excel(excel_file)

        # Convert DataFrame columns to string type
        string_columns = ['Company', 'Job Title', 'Location']
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        matches = []
        search_term_lower = search_term.lower().strip()
        workbook = load_workbook(filename=excel_file)

        def applied_date(row):
            raw = row.get('Applied Date', '')
            raw = '' if str(raw).strip() == 'nan' else raw
            return raw

        for index, row in df.iterrows():
            # Check company name match
            if is_company_match(search_term, row['Company']):
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row),
                    'result': get_result_status(workbook, index + 2),
                })
            # Check exact job title match
            elif search_term_lower in row['Job Title'].lower():
                matches.append({
                    'Company': row['Company'],
                    'Location': row['Location'],
                    'Job Title': row['Job Title'],
                    'Applied Date': applied_date(row),
                    'result': get_result_status(workbook, index + 2),
                })

        workbook.close()
        return matches

    except Exception as e:
        print(f"Error reading file: {str(e)}")
        return []


def parse_markdown_table(markdown_table_string):
    """
    Parses a Markdown table string and returns a list of dictionaries,
    where each dictionary represents a row in the table.
    """
    lines = markdown_table_string.strip().split('\n')
    if len(lines) < 3:
        return []  # Not a valid Markdown table

    # Extract headers
    headers = [header.strip() for header in lines[0].split('|') if header.strip()]

    # Extract rows
    data = []
    for line in lines[2:]:
        values = [value.strip() for value in line.split('|') if value.strip()]
        if len(values) == len(headers):
            data.append(dict(zip(headers, values)))
    return data


def append_data_to_excel(excel_file, data):
    """
    Appends a list of dictionaries to the end of the Excel file using openpyxl.
    Automatically determines the column letters for all fields.
    Adds missing columns if they don't exist.
    """
    try:
        # Load the existing Excel file
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Find the last row with data (or the header row if the sheet is empty)
        last_row = sheet.max_row if sheet.max_row > 1 else 1

        # Get the headers and their column indices
        headers = {cell.value: cell.column for cell in sheet[1]}

        # List of all possible columns
        all_columns = ALL_FIELDS

        # Add any missing columns
        next_column = len(headers) + 1
        for column in all_columns:
            if column not in headers:
                headers[column] = next_column
                sheet.cell(row=1, column=next_column, value=column)
                next_column += 1

        # Append the new data directly to the last row
        for row_data in data:
            last_row += 1  # Move to the next row for each new record

            # Write data to the corresponding columns
            for column in all_columns:
                if column != 'Result':  # Skip Result column as it's handled separately
                    sheet.cell(row=last_row, column=headers[column],
                               value=row_data.get(column, ''))

        # Save the updated workbook
        workbook.save(filename=excel_file)

    except Exception as e:
        print(f"Error appending data to Excel: {str(e)}")


def signal_handler(sig, frame):
    print('\nExiting ...')
    sys.exit(0)


def delete_last_row(excel_file):
    """
    Deletes the last row in the Excel file after user confirmation.
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        last_row = sheet.max_row

        if last_row <= 1:
            print(f"{RED}[*] No data to delete.{RESET}")
            workbook.close()
            return

        # Display the last row's data for confirmation
        last_row_data = [cell.value for cell in sheet[last_row]]
        headers = [cell.value for cell in sheet[1]]
        print("[*] Last row data:")
        for header, value in zip(headers, last_row_data):
            print(f"{header}: {value}")

        # Ask for user confirmation
        confirm = input("[*] Delete this row? (y/Y to confirm, any other key to cancel): ").lower()
        if confirm == 'y':
            sheet.delete_rows(last_row)
            workbook.save(filename=excel_file)
            print(f"{GREEN}[*] Last row deleted successfully.{RESET}")
        else:
            print(f"{RED}[*] Deletion cancelled.{RESET}")

        workbook.close()

    except Exception as e:
        print(f"{RED}[*] Error deleting last row: {str(e)}{RESET}")


def process_webpage_content(content):
    """
    Process webpage content through OpenAI API and return structured data.
    """

    try:
        response = client.beta.chat.completions.parse(
            model=MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": content}
            ],
            temperature=0,
            response_format=JobInfo
        )

        print(response)

        # Parse the response and convert Job_Title to Job Title
        result = response.choices[0].message.parsed
        return {
            "isValid": result.isValid,
            "Company": result.Company,
            "Location": result.Location,
            "Job Title": result.Job_Title,
            "Code": result.Code,
            "Type": result.Type,
            "Link": result.Link
        }
    except Exception as e:
        print(f"Error processing content through OpenAI: {str(e)}")
        return {"isValid": False}


def start_browser(app_path="/Applications/Microsoft Edge Beta.app/Contents/MacOS/Microsoft Edge Beta", url=list(DOMAIN_KEYWORDS.keys())):
    """
    Open a browser to access the specified URL.
    If app_path is provided, it tries to open the URL using the specified application.
    If app_path is not provided, it uses the default browser to open the URL.

    :param app_path: Path to the specific browser application (e.g., /Applications/Microsoft Edge Beta.app)
    :param url: The URL to be accessed (default is None, and no page will be opened if not provided)
    """
    if not url:
        print("Error: URL is not provided, unable to open the browser.")
        return

    if isinstance(url, list):
        success = True
        for u in url:
            success = start_browser(app_path, u)
            if not success:
                success = False
        return success

    if app_path:
        try:
            # Attempt to open the URL with the specified application
            subprocess.run(["open", "-a", app_path, url], check=True)
            print(f"{GREEN}[*] Successfully opened {url} using {app_path}{RESET}")
            return True
        except subprocess.CalledProcessError as e:
            print(f"{RED}[*] Failed to open the specified application, error: {e}{RESET}")
            return False
        except FileNotFoundError:
            print(f"{RED}[*] The specified application path was not found. Please check the path.{RESET}")
            return False


def save_cookie(cookie_path=COOKIE_PATH):
    # Prompt the user to paste the cookie in Netscape format
    print("\n[*] Please paste the cookie in Netscape format (end with an empty line):")
    netscape_cookie = []
    while True:
        line = input()  # Accept multi-line input
        if not line.strip():  # End input on an empty line
            break
        netscape_cookie.append(line)

    # Write the pasted cookie to the file
    try:
        with open(cookie_path, 'w') as f:
            f.write('\n'.join(netscape_cookie))
        print(f"{GREEN}[*] Cookie successfully saved to {cookie_path}{RESET}")
    except IOError as e:
        print(f"{RED}[*] Failed to save the cookie to {cookie_path}: {e}{RESET}")


def validate_cookie(cookie_path=COOKIE_PATH):
    """
    Validate the cookies in the specified cookie file with the provided domain.

    Parameters:
    - domain (str): The base domain of the target server.
    - cookie_path (str): Path to the cookie file in Netscape format.

    Returns:
    - bool: True if the cookies are valid, False otherwise.
    """

    def parseCookieFile(cookiefile):
        """Parse a cookies.txt file and return a dictionary of key value pairs
        compatible with requests."""

        cookies = {}
        with open(cookiefile, 'r') as fp:
            for line in fp:
                if not re.match(r'^\#', line):
                    lineFields = line.strip().split('\t')
                    cookies[lineFields[5]] = lineFields[6]
        return cookies

    try:
        cookies = parseCookieFile(cookie_path)
    except FileNotFoundError:
        print(f"Cookie file not found at path: {cookie_path}")
        return False
    except Exception as e:
        print(f"Failed to load cookies: {e}")
        return False

    success = True
    for url, keywords in DOMAIN_KEYWORDS.items():
        try:
            response = requests.get(url, cookies=cookies, timeout=3)
            if response.status_code == 200 and all(keyword in response.text.lower() for keyword in keywords):
                print(f"{GREEN}[*] {url} LOGGED IN.{RESET}")
            else:
                print(f"{RED}[*] {url} NOT LOGGED IN.{RESET}")
                success = False

        except requests.RequestException as e:
            print(f"{RED}[*] {url} ERROR: {e}{RESET}")
            success = False

    return success


def fetch_webpage_content(url, cookie_path=COOKIE_PATH):
    """
    Fetch content from a URL and extract the main text content.
    """
    try:
        # Add User-Agent header to mimic a browser request
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36'
        }

        # Parse cookies from the cookie file
        cookies = {}
        with open(cookie_path, 'r') as fp:
            for line in fp:
                if not re.match(r'^\#', line):
                    lineFields = line.strip().split('\t')
                    cookies[lineFields[5]] = lineFields[6]

        response = requests.get(url, headers=headers, cookies=cookies, timeout=8)
        response.raise_for_status()

        return response.text

        # Parse HTML content
        soup = BeautifulSoup(response.text, 'html.parser')

        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()

        # Get text content
        text = soup.get_text(separator='\n')

        # Clean up text
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        cleaned_text = '\n'.join(lines)

        return cleaned_text
    except Exception as e:
        print(f"Error fetching webpage: {str(e)}")
        return None


def check_duplicate_entry(excel_file, new_data):
    """
    Check if the exact same job entry already exists in the Excel file.
    Returns True if a duplicate is found, False otherwise.
    """
    try:
        df = pd.read_excel(excel_file)

        # Convert all columns to string for comparison
        for col in df.columns:
            df[col] = df[col].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')

        # Check each row for exact match
        for _, row in df.iterrows():
            all_fields_match = True
            for field in ['Company', 'Job Title']:
                if field in row and field in new_data:
                    if str(row[field]).strip() != str(new_data[field]).strip():
                        all_fields_match = False
                        break
            if all_fields_match:
                return row
        return None

    except Exception as e:
        print(f"{RED}[*] Error checking for duplicates: {str(e)}{RESET}")
        return None


def handle_webpage_content(content, excel_file):
    """
    Handle webpage content: process it and add to Excel if valid
    """
    # Remove view-source: prefix if present
    content = content.strip()
    if content.startswith('view-source:'):
        content = content[11:]  # Remove 'view-source:' prefix

    # Check if content is a URL
    if content.startswith(('http://', 'https://')):
        print("\n[*] Fetching content from URL...")
        webpage_content = fetch_webpage_content(content)
        if not webpage_content:
            print(f"{RED}[*] Failed to fetch webpage content.{RESET}")
            return
        content = "URL: " + content + "\n" + webpage_content

    # Remove extra blank lines
    cleaned_content = '\n'.join(line for line in content.split('\n') if line.strip())

    # Process through OpenAI
    result = process_webpage_content(cleaned_content)

    # Validate required fields one by one
    if not result.get('isValid', False):
        print(f"\n{RED}[*] Invalid content format.{RESET}")
        return

    for field in REQUIRED_FIELDS:
        if field not in result or not str(result[field]).strip():
            print(f"\n{RED}[*] Could not extract valid information from the content.{RESET}")
            return

    # All validations passed, prepare data for Excel
    from datetime import datetime
    data = {
        'Company': result['Company'],
        'Location': result['Location'],
        'Job Title': result['Job Title'],
        'Code': result.get('Code', ''),  # Optional field
        'Type': result.get('Type', ''),  # Optional field
        'Link': result.get('Link', ''),  # Optional field
        'Applied Date': datetime.now().strftime('%Y-%m-%d'),  # Add current date
    }

    # Check for duplicates
    duplicate_entry = check_duplicate_entry(excel_file, data)
    if duplicate_entry is not None:
        print(f"\n{RED}[*] Warning: This job entry already exists in the Excel file.{RESET}")
        print(f"Duplicate Entry: {duplicate_entry}")
        confirm = input("[*] Add it anyway? (y/Y to confirm, any other key to cancel): ").lower()
        if confirm != 'y':
            print(f"\n{RED}[*] Addition cancelled.{RESET}")
            return

    # Add to Excel
    append_data_to_excel(excel_file, [data])

    # Display result
    print(f"\n{GREEN}[*] Successfully extracted and added to Excel:{RESET}")
    print(f"Company: {data['Company']}")
    print(f"Location: {data['Location']}")
    print(f"Job Title: {data['Job Title']}")
    print(f"Code: {data['Code']}")
    print(f"Type: {data['Type']}")
    print(f"Link: {data['Link']}")


def detect_ending(min_threshold=0.05, max_threshold=0.5):
    """
    Detect double Enter press within threshold seconds
    Returns the entered line and a boolean indicating if double Enter was detected
    """

    def end_char_check(line):
        return line.strip().endswith('>') or line.strip().endswith('```')

    line = input("| ")
    if line:
        if end_char_check(line):
            return line, True
        return line, False

    # First empty line detected, start timing
    start_time = time.time()
    try:
        line = input("| ")
        # If second line is entered within threshold seconds
        if not line and min_threshold <= (time.time() - start_time) <= max_threshold:
            return '', True
        elif end_char_check(line):
            return line, True
        return line, False
    except (EOFError, KeyboardInterrupt):
        return '', True


def is_markdown_table(input_string):
    """
    Checks if the input string is likely a Markdown table.
    This is a heuristic check and may not be 100% accurate.
    """
    lines = input_string.strip().split('\n')
    if len(lines) < 3:
        return False  # Not enough lines for a table

    # Check for at least one | in the header and data lines
    if '|' not in lines[0]:
        return False

    for line in lines[2:]:
        if '|' not in line:
            return False

    return True


def summary(excel_file):
    """
    Print a summary of job applications including:
    - Total number of applications
    - Number of rejections
    - Rejection percentage
    """
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active

        # Get total number of applications (excluding header row)
        total_applications = sheet.max_row - 1

        # Count rejections by checking cell fill color
        rejections = 0
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
            result = get_result_status(workbook, row)
            if result.strip():  # If result has an 'x' mark
                rejections += 1

        # Calculate rejection percentage
        rejection_percentage = (rejections / total_applications * 100) if total_applications > 0 else 0

        # Print summary with color formatting
        print(f"\n{GREEN}[*] Application Summary:{RESET}")
        print(f"Total Applications: {total_applications}")
        print(f"{RED}Rejections: {rejections}{RESET}")
        print(f"Rejection Rate: {rejection_percentage:.1f}%")

        workbook.close()

    except Exception as e:
        print(f"{RED}[*] Error generating summary: {str(e)}{RESET}")


def open_excel_file(excel_file):
    """
    Opens the Excel file using the default application based on the operating system.
    Returns True if successful, False otherwise.
    """
    try:
        import subprocess
        import platform
        system = platform.system()
        if system == 'Darwin':  # macOS
            subprocess.run(['open', excel_file])
        elif system == 'Windows':
            subprocess.run(['start', '', excel_file], shell=True)
        else:  # Linux and other OS
            subprocess.run(['xdg-open', excel_file])
        print(f"{GREEN}[*] Opening Excel file...{RESET}")
        return True
    except Exception as e:
        print(f"{RED}[*] Error opening file: {str(e)}{RESET}")
        return False


def main(excel_file=EXCEL_FILE_PATH):
    # Set up signal handler for SIGINT
    signal.signal(signal.SIGINT, signal_handler)

    # Clear the console
    os.system('cls' if os.name == 'nt' else 'clear')

    # Get cookie from the same directory as the script
    cookie_path = COOKIE_PATH

    print(f"[*] Validating cookie...")
    if not validate_cookie(cookie_path):
        print(f"{RED}[*] Cookie is invalid. It is recommended to update the cookie.{RESET}")

    try:
        while True:
            print("\n" + "-" * 100)
            print(
                "[*] Enter search keyword, paste Markdown table, URL, webpage content (starting with '<' or '```'), "
                "\n'delete' to delete last row, 'cookie' to update cookie, 'summary' to view statistics, "
                "(or 'exit' to quit):")
            user_input_lines = []
            line_count = 0
            is_webpage_content = False
            while line_count < 3:
                line = input("> ").lstrip()

                if line.startswith('http://') or line.startswith('https://') or line.startswith('view-source:'):
                    is_webpage_content = True
                    user_input_lines.append(line)
                    break

                if line.startswith('<') or line.startswith('>') or line.startswith('```'):
                    is_webpage_content = True
                    user_input_lines.append(line)
                    try:
                        while True:
                            line, finished = detect_ending()
                            if finished:
                                raise KeyboardInterrupt
                            user_input_lines.append(line)
                    except (EOFError, KeyboardInterrupt):
                        break

                if not line:
                    break
                if "|" not in line:  # Not a Markdown table
                    user_input_lines.append(line)
                    break
                user_input_lines.append(line)
                line_count += 1

            if is_webpage_content:
                print("|" + "-" * 99)
                print("[*] Webpage content detected. Processing ...")
                content = '\n'.join(user_input_lines)
                handle_webpage_content(content, excel_file)
                continue

            user_input = '\n'.join(user_input_lines).strip()

            if user_input.strip().lower() == 'exit':
                print("[*] Exiting ...")
                break

            if user_input.strip().lower() == 'summary':
                summary(excel_file)
                continue

            if user_input.strip().lower().startswith('open'):
                open_excel_file(excel_file)
                continue

            if user_input.strip().lower() == 'delete':
                try:
                    delete_last_row(excel_file)
                except KeyboardInterrupt:
                    print('\n[*] Deletion cancelled. Send SIGINT again to exit.')

                continue

            if user_input.strip().lower() == 'cookie':
                if not validate_cookie(cookie_path):
                    print(f"{RED}[*] Cookie is invalid. Starting cookie update.{RESET}")
                    start_browser()
                    save_cookie(cookie_path)
                    validate_cookie(cookie_path)
                else:
                    print(f"{GREEN}[*] Cookie is valid.{RESET}")
                continue

            if not user_input:
                print(f"{RED}[*] Search keyword cannot be empty!{RESET}")
                continue

            if is_markdown_table(user_input):
                # Parse the Markdown table
                data = parse_markdown_table(user_input)

                if not data:
                    print(f"{RED}[*] Invalid Markdown table format.{RESET}")
                    continue

                # Append the data to the Excel file
                append_data_to_excel(excel_file, data)
                print(f"{GREEN}[*] New record successfully appended to Excel file.{RESET}")

            else:
                results = search_applications(excel_file, user_input)
                if results:
                    print_results(results, excel_file)
                else:
                    print(f"\n{RED}[*] No matching records found.{RESET}")

    except KeyboardInterrupt:
        print('\n[*] Exiting ...')
        sys.exit(0)


if __name__ == "__main__":
    main()
